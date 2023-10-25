from bs4 import BeautifulSoup as bs
import requests
from openpyxl import Workbook

title = []
author = []
rating = []

wb = Workbook()
ws = wb.active
ws["A1"].value = "Title"
ws["B1"].value = "Author"
ws["C1"].value = "Ratings"


def link(url, pages):
    for i in range(1, pages + 1):
        title = []
        author = []
        rating = []
        try:
            response = requests.get(f"{url}?page={i}")
        except:
            print("Tracked All The Pages! ")
            break

        html = response.text
        data = bs(html, "html.parser")

        titles = data.find_all("span", itemprop="name", role="heading")
        for j in titles:
            title.append(j.string)

        authors = data.find_all("a", class_="authorName")
        for j in authors:
            author.append(j.string)

        ratings = data.find_all("span", class_="minirating")
        for j in ratings:
            rating.append(j.get_text().replace(" ", "")[0:4])

        for j in range(len(title)):
            try:
                rating_float = float(rating[j].strip())
                if rating_float >= rating_threshold:
                    ws.append([title[j], author[j], rating_float])
            except ValueError:
                print(f"Skipping item with invalid rating: '{rating[j]}'")

        print(f"Successfully extracted page {i}")


if __name__ == "__main__":
    try:
        url = input(
            "Enter the exact link of the book list from  the goodreads website: "
        )
    except:
        print("Enter a Valid Link! ")

    while True:
        try:
            rating_threshold = float(
                input("Enter the minimum ratings you want to extract: ")
            )
            break
        except ValueError:
            print("Enter a valid number")

    while True:
        try:
            pages = int(input("Enter the total number of Web Pages in Your List: "))
            break
        except ValueError:
            print("Please input a numerical value! ")

    print("Extracting the Data..........................")

    link(url, pages)

    wb.save("books.xlsx")
    print("Excel File Saved Succesfully, Check Your Folder For The File")
